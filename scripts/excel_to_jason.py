from __future__ import annotations

import json
import re
import sys
from copy import deepcopy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from zoneinfo import ZoneInfo


ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "source"
DATA_DIR = ROOT / "data"
FULL_JSON_PATH = DATA_DIR / "menu-data-full.json"
LIGHT_JSON_PATH = DATA_DIR / "menu-data.json"

TZ = ZoneInfo("Asia/Singapore")

DATE_COL_START = 3   # Column C
DATE_COL_END = 9     # Column I

SHEET_PATTERN = re.compile(r"^Week\s+[1-5]$", re.IGNORECASE)

BREAKFAST_ROWS = [4, 5, 6, 7, 8]

LUNCH_NOODLE_ROWS = [9]
LUNCH_DAILY_ROWS = [10, 17, 18, 19]
LUNCH_HOME_ROWS = [11, 12, 13, 14, 15, 16]

DINNER_NOODLE_ROWS = [20]
DINNER_DAILY_ROWS = [21, 28, 29, 30]
DINNER_HOME_ROWS = [22, 23, 24, 25, 26, 27]

SUPPER_NOODLE_ROWS = [31]
SUPPER_DAILY_ROWS = [32, 39, 40, 41]
SUPPER_HOME_ROWS = [33, 34, 35, 36, 37, 38]

ALL_RELEVANT_ROWS = sorted(
    set(
        BREAKFAST_ROWS
        + LUNCH_NOODLE_ROWS + LUNCH_DAILY_ROWS + LUNCH_HOME_ROWS
        + DINNER_NOODLE_ROWS + DINNER_DAILY_ROWS + DINNER_HOME_ROWS
        + SUPPER_NOODLE_ROWS + SUPPER_DAILY_ROWS + SUPPER_HOME_ROWS
    )
)


def log(msg: str) -> None:
    print(msg)


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def load_json_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_json_file(path: Path, payload: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, sort_keys=True)
        f.write("\n")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # Trim each line, remove empty lines
    lines = [line.strip() for line in text.split("\n")]
    lines = [line for line in lines if line]

    text = "\n".join(lines)

    # Collapse excessive spaces inside each line
    text = re.sub(r"[ \t]+", " ", text).strip()

    return text


def non_empty_list(values: List[Any]) -> List[str]:
    out: List[str] = []
    for v in values:
        t = normalize_text(v)
        if t:
            out.append(t)
    return out


def format_date_key(dt: datetime) -> str:
    return dt.strftime("%Y%m%d")


def format_date_text(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


def find_latest_excel(source_dir: Path) -> Path:
    candidates = [p for p in source_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"No .xlsx file found in {source_dir}")
    # Pick latest modified file
    return max(candidates, key=lambda p: p.stat().st_mtime)


def get_cell(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def parse_excel_date(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    return None


def build_breakfast(date_text: str, ws, col: int) -> Optional[Dict[str, Any]]:
    lines = non_empty_list([get_cell(ws, r, col) for r in BREAKFAST_ROWS])
    if not lines:
        return None

    return {
        "date_text": date_text,
        "title": "BREAKFAST MENU",
        "lines": lines,
    }


def build_main_meal(
    *,
    date_text: str,
    title: str,
    noodle_rows: List[int],
    daily_rows: List[int],
    home_rows: List[int],
    ws,
    col: int,
) -> Optional[Dict[str, Any]]:
    noodle_special = non_empty_list([get_cell(ws, r, col) for r in noodle_rows])
    daily_special = non_empty_list([get_cell(ws, r, col) for r in daily_rows])
    home_taste = non_empty_list([get_cell(ws, r, col) for r in home_rows])

    if not noodle_special and not daily_special and not home_taste:
        return None

    return {
        "date_text": date_text,
        "title": title,
        "homeTasteTitle": "HOMETASTE DAILY SELECTION",
        "homeTaste": home_taste,
        "dailySpecialTitle": "DAILY SPECIAL",
        "dailySpecial": daily_special,
        "noodleSpecialTitle": "NOODLE SPECIAL",
        "noodleSpecial": noodle_special,
    }


def is_entire_day_blank(ws, col: int) -> bool:
    for row in ALL_RELEVANT_ROWS:
        if normalize_text(get_cell(ws, row, col)):
            return False
    return True


def parse_day_from_sheet(ws, col: int, dt: datetime) -> Optional[Dict[str, Any]]:
    """
    Return parsed day payload if the day contains any menu content.
    Return None if the whole day is blank.
    """
    if is_entire_day_blank(ws, col):
        return None

    date_text = format_date_text(dt)

    day_payload: Dict[str, Any] = {}

    breakfast = build_breakfast(date_text, ws, col)
    if breakfast:
        day_payload["breakfast"] = breakfast

    lunch = build_main_meal(
        date_text=date_text,
        title="LUNCH MENU",
        noodle_rows=LUNCH_NOODLE_ROWS,
        daily_rows=LUNCH_DAILY_ROWS,
        home_rows=LUNCH_HOME_ROWS,
        ws=ws,
        col=col,
    )
    if lunch:
        day_payload["lunch"] = lunch

    dinner = build_main_meal(
        date_text=date_text,
        title="DINNER MENU",
        noodle_rows=DINNER_NOODLE_ROWS,
        daily_rows=DINNER_DAILY_ROWS,
        home_rows=DINNER_HOME_ROWS,
        ws=ws,
        col=col,
    )
    if dinner:
        day_payload["dinner"] = dinner

    supper = build_main_meal(
        date_text=date_text,
        title="SUPPER MENU",
        noodle_rows=SUPPER_NOODLE_ROWS,
        daily_rows=SUPPER_DAILY_ROWS,
        home_rows=SUPPER_HOME_ROWS,
        ws=ws,
        col=col,
    )
    if supper:
        day_payload["supper"] = supper

    return day_payload if day_payload else None


def parse_workbook(excel_path: Path) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, List[str]]]:
    wb = load_workbook(excel_path, data_only=True)

    parsed_days: Dict[str, Dict[str, Any]] = {}
    duplicates: List[str] = []
    skipped_blank: List[str] = []

    for ws in wb.worksheets:
        if not SHEET_PATTERN.match(ws.title.strip()):
            continue

        for col in range(DATE_COL_START, DATE_COL_END + 1):
            raw_date = get_cell(ws, 3, col)
            dt = parse_excel_date(raw_date)
            if not dt:
                continue

            date_key = format_date_key(dt)
            day_payload = parse_day_from_sheet(ws, col, dt)

            if day_payload is None:
                skipped_blank.append(date_key)
                continue

            if date_key in parsed_days:
                duplicates.append(date_key)

            parsed_days[date_key] = day_payload

    meta = {
        "duplicates": sorted(set(duplicates)),
        "skipped_blank": sorted(set(skipped_blank)),
    }
    return parsed_days, meta


def merge_full_json(
    old_full: Dict[str, Any],
    new_days: Dict[str, Dict[str, Any]],
) -> Tuple[Dict[str, Any], List[str], List[str]]:
    """
    Merge new parsed days into old full JSON.

    Rule:
    - If a date appears in new_days, overwrite that whole date.
    - If a date does not appear in new_days, keep old data.
    """
    merged = deepcopy(old_full)
    new_dates: List[str] = []
    updated_dates: List[str] = []

    for date_key, day_payload in new_days.items():
        if date_key in merged:
            merged[date_key] = day_payload
            updated_dates.append(date_key)
        else:
            merged[date_key] = day_payload
            new_dates.append(date_key)

    merged = dict(sorted(merged.items()))
    return merged, sorted(new_dates), sorted(updated_dates)


def build_light_json(full_data: Dict[str, Any], today: datetime) -> Dict[str, Any]:
    """
    Lightweight front-end JSON:
    - yesterday: keep only supper
    - today to today+7: keep all meals if that date exists
    """
    out: Dict[str, Any] = {}

    yesterday_key = format_date_key(today - timedelta(days=1))
    today_key = format_date_key(today)

    # yesterday supper only
    if yesterday_key in full_data and "supper" in full_data[yesterday_key]:
        out[yesterday_key] = {
            "supper": full_data[yesterday_key]["supper"]
        }

    # today + next 7 days => total 8 calendar days including today
    for offset in range(0, 8):
        day_key = format_date_key(today + timedelta(days=offset))
        if day_key in full_data:
            out[day_key] = deepcopy(full_data[day_key])

    return dict(sorted(out.items()))


def print_summary(
    *,
    excel_path: Path,
    new_dates: List[str],
    updated_dates: List[str],
    skipped_blank: List[str],
    duplicates: List[str],
    full_json: Dict[str, Any],
    light_json: Dict[str, Any],
) -> None:
    log("")
    log("=== Menu JSON Update Summary ===")
    log(f"Source Excel : {excel_path.name}")
    log(f"Full dates   : {len(full_json)}")
    log(f"Light dates  : {len(light_json)}")
    log("")

    log("New dates:")
    if new_dates:
        for d in new_dates:
            log(f"  - {d}")
    else:
        log("  (none)")
    log("")

    log("Updated dates:")
    if updated_dates:
        for d in updated_dates:
            log(f"  - {d}")
    else:
        log("  (none)")
    log("")

    log("Skipped (entire day blank, old data kept):")
    if skipped_blank:
        for d in skipped_blank:
            log(f"  - {d}")
    else:
        log("  (none)")
    log("")

    log("Duplicate dates in workbook (later sheet/column wins):")
    if duplicates:
        for d in duplicates:
            log(f"  - {d}")
    else:
        log("  (none)")
    log("")


def main() -> int:
    ensure_dirs()

    try:
        excel_path = find_latest_excel(SOURCE_DIR)
    except Exception as e:
        log(f"ERROR: {e}")
        return 1

    old_full = load_json_file(FULL_JSON_PATH)

    try:
        new_days, meta = parse_workbook(excel_path)
    except Exception as e:
        log(f"ERROR while parsing workbook: {e}")
        return 1

    merged_full, new_dates, updated_dates = merge_full_json(old_full, new_days)

    today_sg = datetime.now(TZ).replace(hour=0, minute=0, second=0, microsecond=0)
    light_json = build_light_json(merged_full, today_sg)

    try:
        save_json_file(FULL_JSON_PATH, merged_full)
        save_json_file(LIGHT_JSON_PATH, light_json)
    except Exception as e:
        log(f"ERROR while writing JSON files: {e}")
        return 1

    print_summary(
        excel_path=excel_path,
        new_dates=new_dates,
        updated_dates=updated_dates,
        skipped_blank=meta["skipped_blank"],
        duplicates=meta["duplicates"],
        full_json=merged_full,
        light_json=light_json,
    )

    return 0


if __name__ == "__main__":
    sys.exit(main())