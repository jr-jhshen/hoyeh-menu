# -*- coding: utf-8 -*-
"""
Hoyeh Cafeteria Excel -> JSON
Compatible with existing menu-data-full.json structure

Outputs:
1. menu-data-full.json  -> all known dates
2. menu-data.json       -> rolling window (yesterday to +30 days)

Expected Excel layout (per sheet Week 1 ~ Week N):
- C2:I2  => Mon .. Sun
- C3:I3  => dates
- A4     => BREAKFAST
- A9     => LUNCH
- A20    => DINNER
- A31    => SUPPER
- A41    => Vegetarian

Breakfast rows:
- 4..8

Lunch rows:
- noodleSpecial: 9
- dailySpecial: 10,17,18,19
- homeTaste: 11..16

Dinner rows:
- noodleSpecial: 20
- dailySpecial: 21,28,29,30
- homeTaste: 22..27

Supper rows:
- noodleSpecial: 31
- dailySpecial: 32,39,40
- homeTaste: 33..38
- vegetarian: 41   (stored separately, not mixed into supper)
"""

import json
import os
import re
import sys
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


FULL_JSON_FILE = "menu-data-full.json"
ROLLING_JSON_FILE = "menu-data.json"

# Rolling window: yesterday to future 30 days
ROLLING_DAYS_BEFORE = 1
ROLLING_DAYS_AFTER = 30

EXPECTED_WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

DATE_HEADER_ROW = 3
WEEKDAY_HEADER_ROW = 2
START_COL = 3   # C
END_COL = 9     # I

REQUIRED_LABELS = {
    "A4": "BREAKFAST",
    "A9": "LUNCH",
    "A20": "DINNER",
    "A31": "SUPPER",
    "A41": "Vegetarian",
}


class ValidationError(Exception):
    pass


def log_info(msg: str):
    print(f"[INFO] {msg}")


def log_warning(msg: str):
    print(f"[WARNING] {msg}")


def log_error(msg: str):
    print(f"[ERROR] {msg}")


def read_json_file(path: str) -> dict:
    if not os.path.exists(path):
        log_info(f"{path} not found. Starting with empty dataset.")
        return {}

    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            raise ValidationError(f"{path} must contain a JSON object.")
        log_info(f"Loaded existing data from {path} ({len(data)} date(s))")
        return data
    except json.JSONDecodeError as e:
        raise ValidationError(f"Failed to parse JSON file '{path}': {e}")
    except Exception as e:
        raise ValidationError(f"Failed to read JSON file '{path}': {e}")


def save_json_file(path: str, data: dict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log_info(f"Saved {path} ({len(data)} date(s))")


def excel_col_letter(col_idx: int) -> str:
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    # normalize multiple spaces inside each line, but preserve line breaks
    lines = []
    for line in text.split("\n"):
        cleaned = re.sub(r"[ \t]+", " ", line).strip()
        if cleaned:
            lines.append(cleaned)
    return "\n".join(lines)


def split_lines(value) -> list[str]:
    """
    Split cell text into lines.
    Used for Vegetarian or cells with internal line breaks.
    Keeps the order, strips numbering like '1.' / '2)' / '3 ' from the beginning.
    """
    text = normalize_text(value)
    if not text:
        return []

    items = []
    for line in text.split("\n"):
        cleaned = re.sub(r"^\s*\d+\s*[\.\)]\s*", "", line).strip()
        cleaned = re.sub(r"^\s*\d+\s+", "", cleaned).strip()
        if cleaned:
            items.append(cleaned)
    return items


def get_cell_text(ws, cell_ref: str) -> str:
    return normalize_text(ws[cell_ref].value)


def get_row_values(ws, row_num: int, col_num: int, warnings: list[str], context: str) -> list[str]:
    """
    Returns non-empty values from a single row + one column position.
    Since each date is one column, this reads only one cell.
    If the cell contains line breaks, preserve as one string for normal menu cells.
    """
    cell_ref = f"{excel_col_letter(col_num)}{row_num}"
    value = normalize_text(ws[cell_ref].value)
    if not value:
        warnings.append(f"{ws.title} {cell_ref}: blank cell in {context}")
        return []
    return [value]


def get_vertical_block(ws, row_start: int, row_end: int, col_num: int, warnings: list[str], context: str) -> list[str]:
    values = []
    for row_num in range(row_start, row_end + 1):
        cell_ref = f"{excel_col_letter(col_num)}{row_num}"
        value = normalize_text(ws[cell_ref].value)
        if value:
            values.append(value)
        else:
            warnings.append(f"{ws.title} {cell_ref}: blank cell in {context}")
    return values


def parse_excel_date(value, sheet_name: str, cell_ref: str) -> date:
    if value is None:
        raise ValidationError(f"Sheet '{sheet_name}' cell {cell_ref}: date is blank.")

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        raise ValidationError(f"Sheet '{sheet_name}' cell {cell_ref}: date is blank.")

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValidationError(f"Sheet '{sheet_name}' cell {cell_ref}: invalid date '{value}'.")


def format_date_text(dt: date) -> str:
    return dt.strftime("%d/%m/%Y")


def format_date_key(dt: date) -> str:
    return dt.strftime("%Y%m%d")


def validate_sheet_structure(ws):
    # Required labels
    for cell_ref, expected in REQUIRED_LABELS.items():
        actual = get_cell_text(ws, cell_ref)
        if actual != expected:
            raise ValidationError(
                f"Sheet '{ws.title}' cell {cell_ref}: expected '{expected}', got '{actual or '[blank]'}'."
            )

    # Weekday headers
    for idx, expected in enumerate(EXPECTED_WEEKDAYS, start=START_COL):
        cell_ref = f"{excel_col_letter(idx)}{WEEKDAY_HEADER_ROW}"
        actual = get_cell_text(ws, cell_ref)
        if actual != expected:
            raise ValidationError(
                f"Sheet '{ws.title}' cell {cell_ref}: expected weekday '{expected}', got '{actual or '[blank]'}'."
            )

    # Dates present and continuous
    dates = []
    for col in range(START_COL, END_COL + 1):
        cell_ref = f"{excel_col_letter(col)}{DATE_HEADER_ROW}"
        dt = parse_excel_date(ws[cell_ref].value, ws.title, cell_ref)
        dates.append(dt)

    for i in range(1, len(dates)):
        if dates[i] != dates[i - 1] + timedelta(days=1):
            prev_ref = f"{excel_col_letter(START_COL + i - 1)}{DATE_HEADER_ROW}"
            curr_ref = f"{excel_col_letter(START_COL + i)}{DATE_HEADER_ROW}"
            raise ValidationError(
                f"Sheet '{ws.title}': date sequence broken between {prev_ref} ({format_date_text(dates[i - 1])}) "
                f"and {curr_ref} ({format_date_text(dates[i])})."
            )


def build_breakfast(ws, col: int, dt: date, warnings: list[str]) -> dict:
    lines = get_vertical_block(
        ws, 4, 8, col, warnings,
        context=f"BREAKFAST {format_date_text(dt)}"
    )
    if not lines:
        log_warning(f"Sheet '{ws.title}' date {format_date_text(dt)}: breakfast is empty.")

    return {
        "date_text": format_date_text(dt),
        "lines": lines,
        "title": "BREAKFAST MENU",
    }


def build_lunch(ws, col: int, dt: date, warnings: list[str]) -> dict:
    noodle = get_row_values(ws, 9, col, warnings, f"LUNCH noodleSpecial {format_date_text(dt)}")
    daily = []
    daily.extend(get_row_values(ws, 10, col, warnings, f"LUNCH dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 17, col, warnings, f"LUNCH dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 18, col, warnings, f"LUNCH dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 19, col, warnings, f"LUNCH dailySpecial {format_date_text(dt)}"))
    home = get_vertical_block(ws, 11, 16, col, warnings, f"LUNCH homeTaste {format_date_text(dt)}")

    return {
        "dailySpecial": daily,
        "dailySpecialTitle": "DAILY SPECIAL",
        "date_text": format_date_text(dt),
        "homeTaste": home,
        "homeTasteTitle": "HOMETASTE DAILY SELECTION",
        "noodleSpecial": noodle,
        "noodleSpecialTitle": "NOODLE SPECIAL",
        "title": "LUNCH MENU",
    }


def build_dinner(ws, col: int, dt: date, warnings: list[str]) -> dict:
    noodle = get_row_values(ws, 20, col, warnings, f"DINNER noodleSpecial {format_date_text(dt)}")
    daily = []
    daily.extend(get_row_values(ws, 21, col, warnings, f"DINNER dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 28, col, warnings, f"DINNER dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 29, col, warnings, f"DINNER dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 30, col, warnings, f"DINNER dailySpecial {format_date_text(dt)}"))
    home = get_vertical_block(ws, 22, 27, col, warnings, f"DINNER homeTaste {format_date_text(dt)}")

    return {
        "dailySpecial": daily,
        "dailySpecialTitle": "DAILY SPECIAL",
        "date_text": format_date_text(dt),
        "homeTaste": home,
        "homeTasteTitle": "HOMETASTE DAILY SELECTION",
        "noodleSpecial": noodle,
        "noodleSpecialTitle": "NOODLE SPECIAL",
        "title": "DINNER MENU",
    }


def build_supper(ws, col: int, dt: date, warnings: list[str]) -> dict:
    noodle = get_row_values(ws, 31, col, warnings, f"SUPPER noodleSpecial {format_date_text(dt)}")
    daily = []
    daily.extend(get_row_values(ws, 32, col, warnings, f"SUPPER dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 39, col, warnings, f"SUPPER dailySpecial {format_date_text(dt)}"))
    daily.extend(get_row_values(ws, 40, col, warnings, f"SUPPER dailySpecial {format_date_text(dt)}"))
    home = get_vertical_block(ws, 33, 38, col, warnings, f"SUPPER homeTaste {format_date_text(dt)}")

    return {
        "dailySpecial": daily,
        "dailySpecialTitle": "DAILY SPECIAL",
        "date_text": format_date_text(dt),
        "homeTaste": home,
        "homeTasteTitle": "HOMETASTE DAILY SELECTION",
        "noodleSpecial": noodle,
        "noodleSpecialTitle": "NOODLE SPECIAL",
        "title": "SUPPER MENU",
    }


def build_vegetarian(ws, col: int, dt: date, warnings: list[str]) -> dict:
    cell_ref = f"{excel_col_letter(col)}41"
    raw_text = normalize_text(ws[cell_ref].value)
    lines = split_lines(ws[cell_ref].value)

    if not raw_text:
        warnings.append(f"{ws.title} {cell_ref}: Vegetarian menu is blank for {format_date_text(dt)}")

    return {
        "date_text": format_date_text(dt),
        "title": "VEGETARIAN MENU",
        "lines": lines,
        "raw_text": raw_text,
    }


def parse_sheet(ws, warnings: list[str]) -> dict:
    validate_sheet_structure(ws)

    parsed = {}
    for col in range(START_COL, END_COL + 1):
        dt = parse_excel_date(ws.cell(DATE_HEADER_ROW, col).value, ws.title, f"{excel_col_letter(col)}{DATE_HEADER_ROW}")
        date_key = format_date_key(dt)

        if date_key in parsed:
            raise ValidationError(f"Sheet '{ws.title}': duplicate date '{date_key}' within sheet.")

        parsed[date_key] = {
            "breakfast": build_breakfast(ws, col, dt, warnings),
            "lunch": build_lunch(ws, col, dt, warnings),
            "dinner": build_dinner(ws, col, dt, warnings),
            "supper": build_supper(ws, col, dt, warnings),
            "vegetarian": build_vegetarian(ws, col, dt, warnings),
        }

    return parsed


def find_excel_file() -> str:
    # Priority 1: env var
    env_path = os.environ.get("MENU_EXCEL_FILE", "").strip()
    if env_path:
        if os.path.exists(env_path):
            return env_path
        raise ValidationError(f"MENU_EXCEL_FILE is set but file not found: {env_path}")

    # Priority 2: menu.xlsx in repo root
    preferred = Path("menu.xlsx")
    if preferred.exists():
        return str(preferred)

    # Priority 3: a single .xlsx file in repo root
    candidates = [
        p for p in Path(".").glob("*.xlsx")
        if not p.name.startswith("~$")
    ]
    if len(candidates) == 1:
        return str(candidates[0])

    if len(candidates) == 0:
        raise ValidationError(
            "No Excel file found. Put 'menu.xlsx' in repo root, or set MENU_EXCEL_FILE."
        )

    names = ", ".join(sorted(p.name for p in candidates))
    raise ValidationError(
        f"Multiple Excel files found in repo root: {names}. Rename the target file to 'menu.xlsx' "
        f"or set MENU_EXCEL_FILE."
    )


def load_excel_data(excel_path: str) -> tuple[dict, list[str]]:
    warnings = []
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise ValidationError(f"Failed to open Excel file '{excel_path}': {e}")

    if not wb.sheetnames:
        raise ValidationError(f"Excel file '{excel_path}' has no sheets.")

    week_sheets = [name for name in wb.sheetnames if name.strip().lower().startswith("week")]
    if not week_sheets:
        raise ValidationError(f"Excel file '{excel_path}' contains no 'Week' sheets.")

    log_info(f"Reading Excel file: {excel_path}")
    log_info(f"Week sheets found: {', '.join(week_sheets)}")

    all_data = {}

    for sheet_name in week_sheets:
        ws = wb[sheet_name]
        log_info(f"Parsing sheet: {sheet_name}")
        sheet_data = parse_sheet(ws, warnings)

        for date_key, payload in sheet_data.items():
            if date_key in all_data:
                raise ValidationError(
                    f"Duplicate date across sheets: {date_key} already exists before sheet '{sheet_name}'."
                )
            all_data[date_key] = payload

    return dict(sorted(all_data.items())), warnings


def merge_full_data(existing_full: dict, new_data: dict) -> dict:
    merged = deepcopy(existing_full)

    for date_key, day_payload in new_data.items():
        merged[date_key] = day_payload

    return dict(sorted(merged.items()))


def build_rolling_data(full_data: dict) -> dict:
    today = datetime.now().date()
    start_dt = today - timedelta(days=ROLLING_DAYS_BEFORE)
    end_dt = today + timedelta(days=ROLLING_DAYS_AFTER)

    rolling = {}
    for date_key, payload in full_data.items():
        try:
            dt = datetime.strptime(date_key, "%Y%m%d").date()
        except ValueError:
            log_warning(f"Skipping invalid date key in full data: {date_key}")
            continue

        if start_dt <= dt <= end_dt:
            rolling[date_key] = payload

    return dict(sorted(rolling.items()))


def validate_merged_payload(full_data: dict):
    """
    Optional sanity checks after merge.
    """
    for date_key, day_payload in full_data.items():
        if not isinstance(day_payload, dict):
            raise ValidationError(f"Date '{date_key}' payload must be an object.")

        for section in ["breakfast", "lunch", "dinner", "supper"]:
            if section not in day_payload:
                raise ValidationError(f"Date '{date_key}' missing section '{section}'.")

        # vegetarian is now expected too
        if "vegetarian" not in day_payload:
            raise ValidationError(f"Date '{date_key}' missing section 'vegetarian'.")


def main():
    try:
        excel_path = find_excel_file()
        existing_full = read_json_file(FULL_JSON_FILE)
        excel_data, warnings = load_excel_data(excel_path)

        merged_full = merge_full_data(existing_full, excel_data)
        validate_merged_payload(merged_full)

        rolling_data = build_rolling_data(merged_full)

        save_json_file(FULL_JSON_FILE, merged_full)
        save_json_file(ROLLING_JSON_FILE, rolling_data)

        for w in warnings:
            log_warning(w)

        log_info(f"Excel parsed dates: {len(excel_data)}")
        log_info(f"Full dataset dates: {len(merged_full)}")
        log_info(f"Rolling dataset dates: {len(rolling_data)}")
        log_info(f"Warnings: {len(warnings)}")
        log_info("Conversion completed successfully.")

    except ValidationError as e:
        log_error(str(e))
        sys.exit(1)
    except Exception as e:
        log_error(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
